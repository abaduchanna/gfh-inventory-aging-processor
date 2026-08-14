import datetime as _doc_dt
_DOC_YEAR = _doc_dt.date.today().year

f"""
GFH Inventory Aging Processor
==============================
Pick an Inventory Aging Excel export, then:

  PROCESS & EMAIL (local):
    1. Remove SIM/SIM-card rows.
    2. Remove rows with blank Serial 1 (IMEI) or blocked IMEIs.
    3. Keep only devices aged 20+ days.
    4. Merge Arizona D1/D2 into a single Arizona district.
    5. Save a styled .xlsx per district (frozen header, auto-filter, alternating fills).
    6. Email each district file via Outlook Desktop with inline GFH logo signature.

  UPLOAD TO GOOGLE SHEETS:
    1-4 same cleaning as above, but keep devices aged 14+ days.
    5. Upload to Google Sheets ("GFH Inventory Aging") with ONE TAB PER DISTRICT.
    6. Build an "Executive Dashboard" tab highlighting key KPIs and risks.
    7. Every run reconciles the Google Sheet to exactly today's data.

Install once:
    pip install pandas openpyxl gspread gspread-formatting pywin32 tkinterdnd2 pillow

Developed by Abad Umair Channa | Copyright © {date.today().year} | All rights reserved.
"""

import os, sys, re, io, json, base64, tempfile, threading, traceback, queue, argparse, webbrowser
from pathlib import Path
from datetime import datetime

try:
    import tkinter as tk
    from theme_manager import ThemeManager, apply_theme_to_window, get_copyright_year, create_theme_toggle_button
    from tkinter import ttk, filedialog, messagebox, scrolledtext
except ImportError:
    print("tkinter not available."); sys.exit(1)

from header_manager import FixedHeaderManager
from logo_handler import LogoHandler

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False

_MISSING = []
_HAS_GSPREAD = True
_HAS_GSPREAD_FMT = True
_HAS_OUTLOOK = True
_HAS_PIL = True

try:
    import pandas as pd
except ImportError:
    _MISSING.append("pandas")
try:
    import openpyxl
except ImportError:
    _MISSING.append("openpyxl")
try:
    import gspread
    from gspread.exceptions import WorksheetNotFound, SpreadsheetNotFound, APIError
except ImportError:
    _HAS_GSPREAD = False
try:
    import gspread_formatting
    from gspread_formatting import *
except ImportError:
    _HAS_GSPREAD_FMT = False
try:
    import win32com.client as _win32, pythoncom as _pythoncom
except ImportError:
    _HAS_OUTLOOK = False
try:
    from PIL import Image as _PI, ImageTk as _PIT
except ImportError:
    _HAS_PIL = False

# Import openpyxl Border/Side with aliases to avoid conflict with gspread_formatting's Border
from openpyxl.styles import Font, PatternFill, Alignment, Border as XlBorder, Side as XlSide
from openpyxl.utils import get_column_letter

if not _HAS_GSPREAD or not _HAS_GSPREAD_FMT:
    _MISSING.append("gspread + gspread-formatting")
if not _HAS_OUTLOOK:
    _MISSING.append("pywin32")
if not _HAS_PIL:
    _MISSING.append("pillow")

# Show error only for REQUIRED packages (pandas, openpyxl)
_required_missing = [p for p in ("pandas", "openpyxl") if p in _MISSING]
if _required_missing:
    _r = tk.Tk(); _r.withdraw()
    messagebox.showerror(
        "GFH Inventory Aging Processor - missing packages",
        "Required packages not installed:\n\n"
        + "\n".join(f"  - {p}" for p in _required_missing)
        + "\n\nInstall and reopen:\n\npip install " + " ".join(_required_missing),
    )
    sys.exit(1)

# Warn about optional packages
_optional_missing = [p for p in _MISSING if p not in ("pandas", "openpyxl")]
if _optional_missing:
    _r = tk.Tk(); _r.withdraw()
    messagebox.showwarning(
        "GFH Inventory Aging Processor",
        "Optional packages not installed (some features disabled):\n\n"
        + "\n".join(f"  - {p}" for p in _optional_missing)
        + "\n\nInstall with: pip install pandas openpyxl gspread gspread-formatting pywin32 tkinterdnd2 pillow",
    )

# ==========================================================
# CONFIGURATION
# ==========================================================
APP_TITLE = "GFH Inventory Aging Processor"
ICON_ICO_NAME = "gfh_icon.ico"      # used for taskbar + titlebar (Windows .ico)
LOGO_PNG_NAME = "GFH_Telecom_Logo.png"        # used in the header (resized at runtime via PIL)

NAVY       = "#090d26"
EMBEDDED_LOGO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_logo_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_logo_b64.txt"), "r").read().strip()
EMBEDDED_ICON_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_icon_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_icon_b64.txt"), "r").read().strip()

NAVY_DARK  = "#05060f"
RED        = "#f0541c"
RED_DARK   = "#c01820"
LIGHT      = "#f6f7fb"
WHITE      = "#ffffff"
COLOR_BG   = "#f6f7fb"
COLOR_TEXT  = "#1c1c1c"
COLOR_LOG_BG = "#0d0d1f"
COLOR_LOG_TEXT = "#d7e3f0"

COPYRIGHT_TEXT = f"Developed by Abad Umair Channa | Copyright © {date.today().year} | All rights reserved."

# ── Local Excel + Email config ──
MIN_AGE       = 20
KEEP_COLUMNS  = ["District","Store","Product Desc Full","Serial 1","Age in Company"]
BLOCKED_IMEIS = {"358975210745726","350776860110726","358975210799012",
                 "358975210797339","354709280259373","358975210792793",""," "}

CONTACTS_CONFIG_FILE = "gfh_aging_contacts.json"

# Used only to seed gfh_aging_contacts.json the very first time it's created.
# Everything here is editable (and district rows are addable/removable) from
# the app's Settings dialog afterward — this is just the starting point.
_DEFAULT_SENDER = {
    "sender_name": "Abad Umair Channa",
    "sender_email": "[REDACTED]",
    "sender_title": "Senior Inventory & Supply Chain Analyst",
    "sender_mobile": "346-385-9107",
    "cc_name": "",
    "cc_email": "",
}
_DEFAULT_DISTRICTS = {
    "Arizona":       {"name": "Mohammad Farhan Mohiuddin", "email": "[REDACTED]"},
    "Houston":       {"name": "Muhammad Hamza",            "email": "[REDACTED]"},
    "Louisiana":     {"name": "Asif Khan",                 "email": "[REDACTED]"},
    "Colorado West": {"name": "Raiyan Baig",               "email": "[REDACTED]"},
    "Colorado East": {"name": "Shehriyar Ali",             "email": "[REDACTED]"},
    "Tennessee":     {"name": "Ahmed Siraj",                "email": "[REDACTED]"},
}


def _contacts_config_path() -> str:
    return os.path.join(get_app_dir(), CONTACTS_CONFIG_FILE)


def load_contacts_config() -> dict:
    """Load sender/CC/per-district contacts from the JSON config file.
    On first run (no config file yet) this is seeded with the existing
    defaults so nothing looks stripped. After that, the JSON file is the
    single source of truth — districts can be freely added, renamed, or
    removed from the Settings dialog and will persist here, not just the
    original fixed six."""
    path = _contacts_config_path()
    if not os.path.exists(path):
        cfg = dict(_DEFAULT_SENDER)
        cfg["districts"] = {d: dict(v) for d, v in _DEFAULT_DISTRICTS.items()}
        save_contacts_config(cfg)
        return cfg

    cfg = {
        "sender_name": "", "sender_email": "", "sender_title": "", "sender_mobile": "",
        "cc_name": "", "cc_email": "", "districts": {},
    }
    try:
        with open(path, "r", encoding="utf-8") as f:
            saved = json.load(f)
        for k in ("sender_name", "sender_email", "sender_title", "sender_mobile", "cc_name", "cc_email"):
            cfg[k] = saved.get(k, "")
        # Districts are whatever is actually saved — fully dynamic, not
        # restricted to any fixed list, so newly-added districts persist.
        saved_districts = saved.get("districts", {})
        for d, entry in saved_districts.items():
            cfg["districts"][d] = {
                "name": entry.get("name", ""),
                "email": entry.get("email", ""),
            }
    except Exception:
        pass
    return cfg


def save_contacts_config(cfg: dict) -> None:
    path = _contacts_config_path()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


def contacts_config_to_legacy(cfg: dict):
    """Convert the config dict into the SENDER_EMAIL / CC_EMAIL / DISTRICT_EMAILS
    shapes the rest of the pipeline already expects."""
    sender_email = cfg.get("sender_email", "") or ""
    cc_name = cfg.get("cc_name", "")
    cc_email = cfg.get("cc_email", "")
    cc_full = f"{cc_name} <{cc_email}>" if cc_name and cc_email else (cc_email or "")
    district_emails = {}
    for d, entry in cfg.get("districts", {}).items():
        name = entry.get("name", "")
        email = entry.get("email", "")
        if not email:
            continue
        district_emails[d] = f"{name} <{email}>" if name else email
    return sender_email, cc_full, district_emails

# ── Google Sheets config ──
GOOGLE_SHEET_NAME   = "GFH Inventory Aging"
DASHBOARD_TAB_NAME  = "Executive Dashboard"
CREDENTIALS_FILE    = "credentials.json"
AUTHORIZED_USER_FILE = "authorized_user.json"
AGE_THRESHOLD_DAYS  = 14

HEADER_BG = {"red": 0x16 / 255, "green": 0x16 / 255, "blue": 0x32 / 255}
HEADER_FG = {"red": 1, "green": 1, "blue": 1}

COLUMN_CANDIDATES = {
    "District": ["District"],
    "Store": ["Store"],
    "Product Description": ["Product Desc Full", "Product Description", "Product Desc Short", "Product Desc"],
    "Serial 1": ["Serial 1", "Serial1", "Serial"],
    "Age in Company": ["Age in Company", "Age In Company"],
    "PO Date": ["PO Received Date", "PO Date", "PO Receipt Date"],
    "Retail Price": ["Retail Price", "Retail"],
}

# ── Embedded images (base64) ──
GFH_LOGO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "gfh_logo_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "gfh_logo_b64.txt"), "r").read().strip()
ICON_ICO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "icon_ico_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "icon_ico_b64.txt"), "r").read().strip()
HEADER_LOGO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "header_logo_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "header_logo_b64.txt"), "r").read().strip()


def get_script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def get_app_dir() -> str:
    """Directory where user-facing files live (credentials.json, authorized_user.json,
    gfh_aging_contacts.json).

    When frozen into a PyInstaller onefile EXE, ``__file__`` points at the temporary
    ``_MEIPASS`` extraction folder, which is wiped on each run and is never where the
    user placed their files. In that case resolve relative to the EXE itself so files
    sitting next to the app are actually found.
    """
    if getattr(sys, "frozen", False) and getattr(sys, "executable", None):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


# ==========================================================
# DATA CLEANING / PROCESSING
# ==========================================================
class PipelineError(Exception):
    pass

def resolve_column(df, logical_name):
    for candidate in COLUMN_CANDIDATES[logical_name]:
        if candidate in df.columns:
            return candidate
    raise PipelineError(
        f'Could not find a "{logical_name}" column. Looked for: '
        + ", ".join(COLUMN_CANDIDATES[logical_name])
    )



def format_date_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%m/%d/%Y")
    return str(value).strip()



def normalize_district(value):
    text = "" if value is None else str(value).strip()
    if re.match(r"(?i)^arizona\b", text):
        return "Arizona"
    return text




def load_and_clean(excel_path, log):
    """Load Excel, resolve columns, remove SIM/blank/blocked rows, normalize districts."""
    log(f"Reading: {excel_path}")
    df = pd.read_excel(excel_path, engine="openpyxl", dtype=str)
    df = df.dropna(how="all")
    log(f"  Loaded {len(df)} rows.")

    col_district = resolve_column(df, "District")
    col_store = resolve_column(df, "Store")
    col_desc = resolve_column(df, "Product Description")
    col_serial = resolve_column(df, "Serial 1")
    col_age = resolve_column(df, "Age in Company")
    col_po = resolve_column(df, "PO Date")
    col_retail = resolve_column(df, "Retail Price")

    # Remove SIM rows
    is_sim = df[col_desc].astype(str).str.contains(r"\bSIM\b|SIM CARD|SIM KIT", case=False, na=False, regex=True)
    sim_count = int(is_sim.sum())
    df = df[~is_sim].copy()
    log(f"  Removed {sim_count} SIM row(s).")

    # Remove blank/NaN Serial 1
    before = len(df)
    df[col_serial] = df[col_serial].astype(str).str.strip()
    df = df[(df[col_serial] != "") & (df[col_serial].str.lower() != "nan")]
    log(f"  Removed {before - len(df)} row(s) with blank Serial 1.")

    # Remove blocked IMEIs
    before2 = len(df)
    df = df[~df[col_serial].isin(BLOCKED_IMEIS)]
    log(f"  Removed {before2 - len(df)} blocked IMEI row(s).")

    # Convert numeric columns
    df[col_age] = pd.to_numeric(df[col_age], errors="coerce")
    df[col_retail] = pd.to_numeric(df[col_retail], errors="coerce")
    df[col_po] = pd.to_datetime(df[col_po], errors="coerce")

    # Normalize district names (merge Arizona sub-districts)
    df["_District"] = df[col_district].apply(normalize_district)
    merged_count = (df[col_district].astype(str).str.strip() != df["_District"]).sum()
    if merged_count:
        log(f"  Merged {merged_count} row(s) from Arizona sub-districts into 'Arizona'.")

    log(f"  {len(df)} rows remain after cleaning.")
    return df, col_store, col_desc, col_serial, col_age, col_po, col_retail

def build_aged_subset(df, col_age, log):
    aged = df[df[col_age] > AGE_THRESHOLD_DAYS].copy()
    log(f"  {len(aged)} of {len(df)} devices are aged over {AGE_THRESHOLD_DAYS} days.")
    return aged



def build_district_tabs(aged, col_store, col_desc, col_serial, col_age, col_po, log):
    tabs = {}
    for district in sorted(aged["_District"].dropna().unique()):
        if not district:
            continue
        group = aged[aged["_District"] == district]
        tab_df = pd.DataFrame(
            {
                "District": district,
                "Store": group[col_store].fillna("").astype(str).str.strip(),
                "Product Description": group[col_desc].fillna("").astype(str).str.strip(),
                "Serial 1": group[col_serial],
                "Age in Company": group[col_age],
                "PO Date": group[col_po].apply(format_date_value),
            }
        ).sort_values(["Store", "Age in Company"], ascending=[True, False]).reset_index(drop=True)
        tabs[district] = tab_df
        log(f"  {district}: {len(tab_df)} aged device row(s)")
    return tabs


# ==========================================================
# GOOGLE SHEETS UPLOAD
# ==========================================================


# ==========================================================
# LOCAL EXCEL + EMAIL
# ==========================================================

def save_district_xlsx(df_d, district, out_dir):
    safe     = re.sub(r'[\\/:*?"<>|]',"_",district).strip()
    out_path = Path(out_dir) / f"Inventory_Aging_District_{safe}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = district[:31]
    headers = list(df_d.columns)
    hf = Font(name="Calibri",bold=True,color="FFFFFF",size=11)
    hfill = PatternFill("solid",fgColor="1F3864")
    ha = Alignment(horizontal="center",vertical="center",wrap_text=True)
    thin = XlSide(style="thin",color="FFFFFF")
    hb = XlBorder(left=thin,right=thin,top=thin,bottom=thin)
    for ci,h in enumerate(headers,1):
        c=ws.cell(row=1,column=ci,value=h); c.font=hf; c.fill=hfill; c.alignment=ha; c.border=hb
    ws.row_dimensions[1].height=28
    df_ = Font(name="Calibri",size=10)
    da = Alignment(vertical="center")
    db = XlBorder(left=XlSide(style="thin",color="CCCCCC"),right=XlSide(style="thin",color="CCCCCC"),
                top=XlSide(style="thin",color="CCCCCC"),bottom=XlSide(style="thin",color="CCCCCC"))
    af = PatternFill("solid",fgColor="EAF2FF")
    col_w={"District":18,"Store":22,"Product Desc Full":45,"Serial 1":20,"Age in Company":16}
    for ri,(_,row) in enumerate(df_d.iterrows(),2):
        fill=af if ri%2==0 else None
        for ci,h in enumerate(headers,1):
            v=row[h]
            if h=="Serial 1":
                cell=ws.cell(row=ri,column=ci,value=str(v) if pd.notna(v) else ""); cell.number_format="@"
            elif h=="Age in Company":
                cell=ws.cell(row=ri,column=ci,value=int(v) if pd.notna(v) else "")
            else:
                cell=ws.cell(row=ri,column=ci,value=v if pd.notna(v) else "")
            cell.font=df_; cell.alignment=da; cell.border=db
            if fill: cell.fill=fill
    for ci,h in enumerate(headers,1):
        ws.column_dimensions[get_column_letter(ci)].width=col_w.get(h,18)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    wb.save(out_path); return out_path

# ─────────────────────────────────────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def build_html_email(district, to_address="", sender_cfg=None):
    """sender_cfg: dict with sender_name/sender_email/sender_title/sender_mobile,
    all editable in Settings. No identity is hardcoded here any more."""
    sender_cfg = sender_cfg or {}
    sender_name   = sender_cfg.get("sender_name", "") or "GFH Telecom"
    sender_email  = sender_cfg.get("sender_email", "")
    sender_title  = sender_cfg.get("sender_title", "")
    sender_mobile = sender_cfg.get("sender_mobile", "")

    today = datetime.now().strftime("%B %d, %Y")
    # Extract first name from "Full Name <email>" or "Full Name"
    dm_name = "Team"
    if to_address:
        raw = to_address.split("<")[0].strip()   # "Full Name"
        if raw:
            dm_name = raw.split()[0]

    email_row = (
        f'<tr><td style="color:#f0541c;font-weight:700;padding-right:6px;font-family:Calibri;">E:</td>'
        f'<td style="font-family:Calibri;"><a href="mailto:{sender_email}" '
        f'style="color:#090d26;text-decoration:none;">{sender_email}</a></td></tr>'
    ) if sender_email else ""
    mobile_row = (
        f'<tr><td style="color:#f0541c;font-weight:700;padding-right:6px;font-family:Calibri;">M:</td>'
        f'<td style="color:#090d26;font-family:Calibri;">{sender_mobile}</td></tr>'
    ) if sender_mobile else ""

    return f"""
<html><body style="font-family:Calibri,Arial,sans-serif;font-size:14px;color:#1a1a1a;">
<p>Dear {dm_name},</p>
<p>Please find attached the <strong>Inventory Aging Report — {district} District</strong> as of {today}.</p>
<p>This report includes all devices with <strong>20+ days</strong> age in company. Kindly review and take necessary action on aging inventory.</p>
<p>Thank you,</p>
<br>
<table style="border-top:2px solid #f0541c;padding-top:12px;">
  <tr>
    <td style="padding-right:16px;vertical-align:top;">
      <img src="cid:gfhlogo" width="170" alt="GFH Telecom">
    </td>
    <td style="border-left:2px solid #f0541c;padding-left:16px;vertical-align:top;font-family:Calibri,Arial,sans-serif;">
      <div style="font-size:16px;font-weight:700;color:#090d26;">{sender_name}</div>
      <div style="font-size:11px;color:#666;margin-top:2px;">{sender_title}</div>
      <table style="margin-top:8px;font-size:12px;border-collapse:collapse;">
        {email_row}
        {mobile_row}
      </table>
    </td>
  </tr>
</table>
</body></html>"""


def send_via_outlook(to, subject, html_body, attachment_path, log, sender_email="", cc_email=""):
    try:
        import win32com.client as win32, pythoncom
        pythoncom.CoInitialize()
    except ImportError:
        log("ERROR: pywin32 not installed. Run: pip install pywin32"); return False
    logo_tmp = None
    try:
        # Decode logo to temp file
        data = base64.b64decode(GFH_LOGO_B64.replace("\n","").replace(" ","").strip())
        logo_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        logo_tmp.write(data); logo_tmp.close(); logo_tmp = logo_tmp.name

        outlook = win32.Dispatch("Outlook.Application")
        mail    = outlook.CreateItem(0)

        # Set sender account, if configured and matching a signed-in Outlook account
        if sender_email:
            for acc in outlook.Session.Accounts:
                if sender_email.lower() in str(acc.SmtpAddress).lower():
                    mail._oleobj_.Invoke(*(64209,0,8,0,acc)); break

        mail.To      = to
        if cc_email:
            mail.CC  = cc_email
        mail.Subject = subject

        # Inline logo attachment (CID)
        att = mail.Attachments.Add(logo_tmp)
        att.PropertyAccessor.SetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x3712001F","gfhlogo")
        att.PropertyAccessor.SetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x7FFF000B",True)

        mail.HTMLBody = html_body

        # District Excel attachment
        if attachment_path and Path(attachment_path).exists():
            mail.Attachments.Add(str(Path(attachment_path).resolve()))

        mail.Send()
        return True
    except Exception as exc:
        log(f"Outlook error: {exc}"); return False
    finally:
        if logo_tmp and os.path.exists(logo_tmp):
            try: os.unlink(logo_tmp)
            except: pass

# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────


# ==========================================================
# GOOGLE SHEETS UPLOAD
# ==========================================================

def authenticate_google(log):
    creds_path = os.path.join(get_app_dir(), CREDENTIALS_FILE)
    auth_user_path = os.path.join(get_app_dir(), AUTHORIZED_USER_FILE)

    if not os.path.exists(creds_path):
        raise PipelineError(
            f"{CREDENTIALS_FILE} not found beside this script.\n"
            f"Expected here:\n{creds_path}\n\n"
            "Place your Google OAuth Desktop credentials.json in the same "
            "folder as this app (same file used by GFH Rebate Google Drive Sync)."
        )

    log("Connecting to Google (a browser window may open to authorize)...")
    return gspread.oauth(
        credentials_filename=creds_path,
        authorized_user_filename=auth_user_path,
    )



def open_or_create_spreadsheet(gc, log):
    try:
        return gc.open(GOOGLE_SHEET_NAME)
    except SpreadsheetNotFound:
        log(f"Google Sheet '{GOOGLE_SHEET_NAME}' not found. Creating it...")
        return gc.create(GOOGLE_SHEET_NAME)



def get_or_create_worksheet(spreadsheet, title, rows, cols):
    try:
        worksheet = spreadsheet.worksheet(title)
    except WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=title, rows=max(rows, 2), cols=max(cols, 1))
    worksheet.resize(rows=max(rows, 2), cols=max(cols, 1))
    return worksheet



def dataframe_to_values(df):
    values = [list(df.columns)]
    for row in df.itertuples(index=False, name=None):
        clean_row = []
        for v in row:
            if pd.isna(v):
                clean_row.append("")
            elif isinstance(v, bool):
                clean_row.append(v)
            elif isinstance(v, (int, float)):
                clean_row.append(v.item() if hasattr(v, "item") else v)
            else:
                clean_row.append(str(v))
        values.append(clean_row)
    return values



def clear_and_upload_values(worksheet, values, value_input_option="RAW"):
    worksheet.clear()
    if not values:
        return
    worksheet.update(values=values, range_name="A1", value_input_option=value_input_option)



def apply_header_formatting(spreadsheet, worksheet, row_count, col_count):
    sheet_id = worksheet.id
    requests = [
        {
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": col_count,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": HEADER_BG,
                        "horizontalAlignment": "CENTER",
                        "textFormat": {"foregroundColor": HEADER_FG, "bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
            }
        },
        {
            "autoResizeDimensions": {
                "dimensions": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": col_count}
            }
        },
    ]
    if row_count > 1 and col_count > 0:
        requests.append(
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": row_count,
                            "startColumnIndex": 0,
                            "endColumnIndex": col_count,
                        }
                    }
                }
            }
        )
    spreadsheet.batch_update({"requests": requests})



def sanitize_tab_name(name):
    cleaned = re.sub(r"[\[\]\*\?/\\:]", "-", str(name)).strip()
    return (cleaned or "Unknown")[:100]



def build_executive_dashboard(sh, df_aged):
    sheet_name = DASHBOARD_TAB_NAME
    
    try:
        ws = sh.worksheet(sheet_name)
        ws.clear()
        try:
            ws.clear_basic_filter()
        except:
            pass
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_name, rows="100", cols="15")

    sh.reorder_worksheets([ws] + [s for s in sh.worksheets() if s.title != sheet_name])

    col_retail = resolve_column(df_aged, "Retail Price")
    
    total_value = float(pd.to_numeric(df_aged[col_retail], errors="coerce").sum())
    total_devices = len(df_aged)
    
    district_risk = df_aged.groupby('_District')[col_retail].sum().reset_index()
    district_risk = district_risk.sort_values(col_retail, ascending=False)
    top_district = district_risk.iloc[0]['_District'] if not district_risk.empty else "N/A"
    top_district_val = float(district_risk.iloc[0][col_retail]) if not district_risk.empty else 0

    col_store = resolve_column(df_aged, "Store")
    store_risk = df_aged.groupby(col_store)[col_retail].sum().reset_index()
    store_risk = store_risk.sort_values(col_retail, ascending=False)
    top_store = store_risk.iloc[0][col_store] if not store_risk.empty else "N/A"
    top_store_val = float(store_risk.iloc[0][col_retail]) if not store_risk.empty else 0

    col_serial = resolve_column(df_aged, "Serial 1")

    dashboard_data = [
        ["GFH TELECOM - AGED INVENTORY DASHBOARD (>14 Days)", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["TOTAL VALUE AT RISK", "", "TOTAL AGED DEVICES", "", "HIGHEST RISK DISTRICT", "", "HIGHEST RISK STORE"],
        [f"${total_value:,.2f}", "", f"{total_devices}", "", f"{top_district} (${top_district_val:,.2f})", "", f"{top_store} (${top_store_val:,.2f})"],
        ["", "", "", "", "", "", ""],
        ["", "", "", "", "", "", ""],
        ["DISTRICT BREAKDOWN", "", "", "", "TOP 15 STORES AT RISK", "", ""],
        ["District", "Aged Devices", "Total Value ($)", "", "Store Name", "District", "Total Value ($)"]
    ]

    district_summary = df_aged.groupby('_District').agg(
        Devices=(col_serial, 'count'),
        Total_Value=(col_retail, 'sum')
    ).reset_index().sort_values('Total_Value', ascending=False)

    top_15_stores = df_aged.groupby([col_store, '_District']).agg(
        Total_Value=(col_retail, 'sum')
    ).reset_index().sort_values('Total_Value', ascending=False).head(15)

    max_rows = max(len(district_summary), len(top_15_stores))

    for i in range(max_rows):
        row = []
        if i < len(district_summary):
            row.extend([
                district_summary.iloc[i]['_District'],
                int(district_summary.iloc[i]['Devices']),
                f"${float(district_summary.iloc[i]['Total_Value']):,.2f}"
            ])
        else:
            row.extend(["", "", ""])
        
        row.append("")

        if i < len(top_15_stores):
            row.extend([
                top_15_stores.iloc[i][col_store],
                top_15_stores.iloc[i]['_District'],
                f"${float(top_15_stores.iloc[i]['Total_Value']):,.2f}"
            ])
        else:
            row.extend(["", "", ""])

        dashboard_data.append(row)

    ws.update(range_name='A1', values=dashboard_data)

    header_fmt = cellFormat(
        backgroundColor=color(0.08, 0.08, 0.2),
        textFormat=textFormat(bold=True, foregroundColor=color(1, 1, 1), fontSize=14),
        horizontalAlignment='CENTER'
    )
    kpi_title_fmt = cellFormat(
        backgroundColor=color(0.9, 0.9, 0.9),
        textFormat=textFormat(bold=True, fontSize=10, foregroundColor=color(0.3, 0.3, 0.3)),
        horizontalAlignment='CENTER'
    )
    kpi_value_fmt = cellFormat(
        textFormat=textFormat(bold=True, fontSize=16, foregroundColor=color(0.8, 0, 0)),
        horizontalAlignment='CENTER'
    )
    table_header_fmt = cellFormat(
        backgroundColor=color(0.2, 0.2, 0.2),
        textFormat=textFormat(bold=True, foregroundColor=color(1, 1, 1)),
        horizontalAlignment='CENTER'
    )

    formats = [
        ('A1:G1', header_fmt),
        ('A3:B3', kpi_title_fmt), ('C3:D3', kpi_title_fmt), 
        ('E3:F3', kpi_title_fmt), ('G3:H3', kpi_title_fmt),
        ('A4:B4', kpi_value_fmt), ('C4:D4', kpi_value_fmt), 
        ('E4:F4', kpi_value_fmt), ('G4:H4', kpi_value_fmt),
        ('A7:C7', header_fmt), ('E7:G7', header_fmt),
        ('A8:C8', table_header_fmt), ('E8:G8', table_header_fmt)
    ]
    
    format_cell_ranges(ws, formats)

    ws.merge_cells('A1:H1') 
    ws.merge_cells('A3:B3') 
    ws.merge_cells('A4:B4') 
    ws.merge_cells('C3:D3') 
    ws.merge_cells('C4:D4') 
    ws.merge_cells('E3:F3') 
    ws.merge_cells('E4:F4') 
    ws.merge_cells('G3:H3') 
    ws.merge_cells('G4:H4') 

    ws.merge_cells('A7:C7')
    ws.merge_cells('E7:G7')

    set_column_width(ws, 'A', 180)
    set_column_width(ws, 'B', 100)
    set_column_width(ws, 'C', 120)
    set_column_width(ws, 'D', 40)
    set_column_width(ws, 'E', 250)
    set_column_width(ws, 'F', 180)
    set_column_width(ws, 'G', 120)

    return True



def upload_to_google_sheets(district_tabs, aged_df, log):
    gc = authenticate_google(log)
    spreadsheet = open_or_create_spreadsheet(gc, log)
    log(f"Google Sheet ready: {spreadsheet.url}")

    target_tab_names = {sanitize_tab_name(d) for d in district_tabs.keys()}
    target_tab_names.add(DASHBOARD_TAB_NAME)

    log("Building Executive Dashboard...")
    try:
        build_executive_dashboard(spreadsheet, aged_df)
    except Exception as e:
        log(f"Warning: Failed to build dashboard: {e}")

    district_worksheets = []
    for district, tab_df in district_tabs.items():
        tab_name = sanitize_tab_name(district)
        log(f"Uploading '{tab_name}' tab ({len(tab_df)} rows)...")
        values = dataframe_to_values(tab_df)
        ws = get_or_create_worksheet(spreadsheet, tab_name, rows=len(values), cols=len(values[0]))
        clear_and_upload_values(ws, values)
        apply_header_formatting(spreadsheet, ws, len(values), len(values[0]))
        district_worksheets.append(ws)

    try:
        dashboard_ws = spreadsheet.worksheet(DASHBOARD_TAB_NAME)
        ordered = [dashboard_ws] + sorted(district_worksheets, key=lambda w: w.title)
        spreadsheet.reorder_worksheets(ordered)
    except Exception:
        pass

    removed = 0
    for ws in spreadsheet.worksheets():
        if ws.title not in target_tab_names:
            spreadsheet.del_worksheet(ws)
            removed += 1
    if removed:
        log(f"Removed {removed} leftover tab(s) from previous runs.")

    log("Upload complete.")
    return spreadsheet.url




# ==========================================================
# PIPELINE ORCHESTRATORS
# ==========================================================

def run_local_pipeline(excel_path, log, send_email=True):
    """Process Excel locally: save district .xlsx files, optionally email via Outlook."""
    log("=" * 60)
    log("GFH INVENTORY AGING: PROCESS & EMAIL")
    log("=" * 60)

    df, col_store, col_desc, col_serial, col_age, col_po, col_retail = load_and_clean(excel_path, log)

    aged = df[df[col_age] >= MIN_AGE].copy()
    log(f"  {len(aged)} of {len(df)} devices are aged {MIN_AGE}+ days.")

    districts = sorted(aged["_District"].dropna().unique())
    log(f"  Districts: {', '.join(districts)}")
    log("")

    saved = []
    for i, dist in enumerate(districts, 1):
        df_d = aged[aged["_District"] == dist][KEEP_COLUMNS].copy()
        df_d = df_d.sort_values(by="Store", key=lambda s: s.str.lower(),
                                kind="mergesort", ignore_index=True)
        log(f"[{i}/{len(districts)}] Saving: {dist}  ({len(df_d)} rows)")
        path = save_district_xlsx(df_d, dist, Path(excel_path).parent)
        log(f"  Saved: {path.name}")
        saved.append((dist, path))

    if send_email and _HAS_OUTLOOK:
        cfg = load_contacts_config()
        sender_email, cc_full, district_emails = contacts_config_to_legacy(cfg)
        log("")
        log("Sending emails via Outlook Desktop...")
        for dist, fp in saved:
            to = district_emails.get(dist, "")
            if not to:
                log(f"  SKIP {dist} - no email configured (set it in Settings)"); continue
            log(f"  -> {dist}  >  {to}")
            subj = f"GFH Telecom - Inventory Aging Report: {dist} District"
            body = build_html_email(dist, to, sender_cfg=cfg)
            ok = send_via_outlook(to, subj, body, fp, log, sender_email=sender_email, cc_email=cc_full)
            log(f"    {'Sent' if ok else 'FAILED'}")

    log("")
    log(f"COMPLETE - {len(saved)} district file(s) processed")
    return saved


def run_google_pipeline(excel_path, log):
    """Process Excel and upload to Google Sheets."""
    if not _HAS_GSPREAD or not _HAS_GSPREAD_FMT:
        raise PipelineError(
            "Google Sheets upload requires gspread + gspread-formatting.\n"
            "Install with: pip install gspread gspread-formatting")

    log("=" * 60)
    log("GFH INVENTORY AGING: UPLOAD TO GOOGLE SHEETS")
    log("=" * 60)

    df, col_store, col_desc, col_serial, col_age, col_po, col_retail = load_and_clean(excel_path, log)

    log("")
    log(f"Filtering to devices aged over {AGE_THRESHOLD_DAYS} days...")
    aged = build_aged_subset(df, col_age, log)

    log("")
    log("Building one tab per district...")
    district_tabs = build_district_tabs(aged, col_store, col_desc, col_serial, col_age, col_po, log)

    total_count = len(aged)
    total_loss = float(pd.to_numeric(aged[col_retail], errors="coerce").sum())

    log("")
    sheet_url = upload_to_google_sheets(district_tabs, aged, log)

    log("")
    log("=" * 60)
    log("SUCCESS")
    log(f"Districts uploaded: {len(district_tabs)}")
    log(f"Devices aged >{AGE_THRESHOLD_DAYS} days: {total_count}")
    log(f"Total potential loss: ${total_loss:,.2f}")
    log(f"Google Sheet: {sheet_url}")
    log("=" * 60)

    return sheet_url


# ==========================================================
# GUI
# ==========================================================



def _extract_embedded_icon(b64, filename):
    """Decode an embedded base64 icon to a temp file; return path or None."""
    try:
        if not b64:
            return None
        import base64 as _b64, tempfile, os
        target = os.path.join(tempfile.gettempdir(), filename)
        with open(target, "wb") as fh:
            fh.write(_b64.b64decode(b64))
        return target if os.path.isfile(target) else None
    except Exception:
        return None

def _set_window_icon(root):
    """Set taskbar + titlebar icon from embedded base64 ICO."""
    import base64, tempfile, atexit, os, sys

    # 1. Try sys._MEIPASS (PyInstaller onefile extraction dir)
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        ico_path = os.path.join(meipass, "gfh_icon.ico")
        if os.path.exists(ico_path):
            try:
                root.iconbitmap(default=False, bitmap=ico_path)
                root.iconbitmap(ico_path)
                return
            except Exception:
                pass

    # 2. Try next to the exe/script
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    ico_path = os.path.join(base_dir, "gfh_icon.ico")
    if os.path.exists(ico_path):
        try:
            root.iconbitmap(default=False, bitmap=ico_path)
            root.iconbitmap(ico_path)
            return
        except Exception:
            pass

    # 3. Decode EMBEDDED_ICON_B64 to %TEMP% (no spaces, always writable)
    try:
        data = base64.b64decode(EMBEDDED_ICON_B64.strip())
        tmp_dir = os.environ.get("TEMP", tempfile.gettempdir())
        ico_path = os.path.join(tmp_dir, "gfh_app_icon.ico")
        with open(ico_path, "wb") as f:
            f.write(data)
        root.iconbitmap(default=False, bitmap=ico_path)
        root.iconbitmap(ico_path)
        return
    except Exception:
        pass


class SettingsDialog(tk.Toplevel):
    """Lets the user enter the sender identity, CC, and a name+email per
    district. Districts are fully dynamic here — add, rename, or remove
    rows freely; whatever's in the dialog on Save becomes the new district
    list, persisted to CONTACTS_CONFIG_FILE (gfh_aging_contacts.json)."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Settings — Sender & District Contacts")
        self.configure(bg=LIGHT)
        self.resizable(False, True)
        self.transient(parent)
        self.grab_set()

        cfg = load_contacts_config()
        self._vars = {}
        self._dist_rows = []  # list of (row_frame, district_name_var, name_var, email_var)

        pad = {"padx": 8, "pady": 4}

        sender_frame = tk.LabelFrame(self, text="Sender identity (used as email signature)",
                                      bg=LIGHT, fg=NAVY, font=("Calibri", 10, "bold"))
        sender_frame.pack(fill="x", padx=12, pady=(12, 6))

        def add_row(frame, row, label, key, default=""):
            tk.Label(frame, text=label, bg=LIGHT, fg=NAVY,
                     font=("Calibri", 9)).grid(row=row, column=0, sticky="w", **pad)
            v = tk.StringVar(value=cfg.get(key, default))
            tk.Entry(frame, textvariable=v, width=40, font=("Calibri", 9)
                     ).grid(row=row, column=1, sticky="ew", **pad)
            self._vars[key] = v

        add_row(sender_frame, 0, "Your name:", "sender_name")
        add_row(sender_frame, 1, "Your email:", "sender_email")
        add_row(sender_frame, 2, "Your title:", "sender_title")
        add_row(sender_frame, 3, "Your mobile:", "sender_mobile")
        add_row(sender_frame, 4, "CC name:", "cc_name")
        add_row(sender_frame, 5, "CC email:", "cc_email")
        sender_frame.columnconfigure(1, weight=1)

        dist_outer = tk.LabelFrame(self, text="District contacts (add/remove districts freely)",
                                    bg=LIGHT, fg=NAVY, font=("Calibri", 10, "bold"))
        dist_outer.pack(fill="both", expand=True, padx=12, pady=(6, 6))

        hdr = tk.Frame(dist_outer, bg=LIGHT); hdr.pack(fill="x", padx=8, pady=(6, 0))
        tk.Label(hdr, text="District", bg=LIGHT, fg="#666", font=("Calibri", 8), width=16, anchor="w").pack(side="left")
        tk.Label(hdr, text="Contact name", bg=LIGHT, fg="#666", font=("Calibri", 8), width=20, anchor="w").pack(side="left")
        tk.Label(hdr, text="Email", bg=LIGHT, fg="#666", font=("Calibri", 8), width=28, anchor="w").pack(side="left")

        self._dist_container = tk.Frame(dist_outer, bg=LIGHT)
        self._dist_container.pack(fill="both", expand=True, padx=8, pady=(2, 6))

        for d, entry in cfg["districts"].items():
            self._add_district_row(d, entry.get("name", ""), entry.get("email", ""))

        ttk.Button(dist_outer, text="+ Add District", style="Browse.TButton",
                   command=lambda: self._add_district_row("", "", "")
                   ).pack(anchor="w", padx=8, pady=(0, 8))

        btn_row = tk.Frame(self, bg=LIGHT)
        btn_row.pack(fill="x", padx=12, pady=(6, 12))
        ttk.Button(btn_row, text="Save", style="Run.TButton",
                   command=self._save).pack(side="right", padx=(6, 0))
        ttk.Button(btn_row, text="Cancel", style="Browse.TButton",
                   command=self.destroy).pack(side="right")

    def _add_district_row(self, district_name, contact_name, contact_email):
        row = tk.Frame(self._dist_container, bg=LIGHT)
        row.pack(fill="x", pady=2)

        dv = tk.StringVar(value=district_name)
        nv = tk.StringVar(value=contact_name)
        ev = tk.StringVar(value=contact_email)

        tk.Entry(row, textvariable=dv, width=16, font=("Calibri", 9)).pack(side="left", padx=(0, 4))
        tk.Entry(row, textvariable=nv, width=20, font=("Calibri", 9)).pack(side="left", padx=(0, 4))
        tk.Entry(row, textvariable=ev, width=28, font=("Calibri", 9)).pack(side="left", padx=(0, 4))

        entry_tuple = (row, dv, nv, ev)

        def remove():
            row.destroy()
            self._dist_rows.remove(entry_tuple)

        tk.Button(row, text="✕", command=remove, bg=LIGHT, fg="#f0541c",
                  font=("Calibri", 9, "bold"), bd=0, cursor="hand2").pack(side="left")

        self._dist_rows.append(entry_tuple)

    def _save(self):
        cfg = {k: v.get().strip() for k, v in self._vars.items()}
        districts = {}
        for _row, dv, nv, ev in self._dist_rows:
            d = dv.get().strip()
            if not d:
                continue  # skip rows where the district name was left blank
            districts[d] = {"name": nv.get().strip(), "email": ev.get().strip()}
        cfg["districts"] = districts
        save_contacts_config(cfg)
        messagebox.showinfo("Saved", "Contacts saved to gfh_aging_contacts.json", parent=self)
        self.destroy()


class App:
    def __init__(self, root):
        self.root = root
        self.file_path = tk.StringVar()
        self._q = queue.Queue()
        self._running = False
        self.sheet_url_cache = None
        self._wordmark_img = None

        root.title(APP_TITLE)
        # Dynamic screen resolution support: size to 90% of the screen and
        # center it (DPI-aware), then stay a normal resizable top-level so
        # Windows Snap (50% left/right, corners, Win+arrow) keeps working.
        self._apply_dynamic_geometry()
        root.configure(bg=LIGHT)
        # Brute-force taskbar icon: set AppUserModelID so Windows taskbar
        # shows our icon instead of the generic Python/PyInstaller icon
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                "GFHTelecom.App")
        except Exception:
            pass

        _set_window_icon(root)
        root.protocol("WM_DELETE_WINDOW", root.destroy)
        self.theme_manager = ThemeManager("GFH Inventory Aging Processor", app_name="gfh-inventory-aging-processor")
        self._styles(); self._header(); self._body(); self._copyright_bar()
        apply_theme_to_window(self.root, self.theme_manager)
        self._poll()

    def _apply_dynamic_geometry(self) -> None:
        """Size the window to 90% of the screen and center it.

        Works on any laptop/monitor/PC (1080p, 1440p, 2K, 4K) and respects
        Windows DPI scaling (run after _enable_dpi_awareness()). The window
        stays resizable so Windows Snap gestures keep working — it centers
        on launch, then snaps normally to 50% left/right, corners or via
        Win+arrow shortcuts.
        """
        try:
            root = self.root
            root.update_idletasks()
            sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
            w = max(640, min(int(sw * 0.90), sw - 20))
            h = max(480, min(int(sh * 0.90), sh - 40))
            x = max(0, (sw - w) // 2)
            y = max(0, (sh - h) // 2)
            root.geometry(f"{w}x{h}+{x}+{y}")
            root.minsize(min(660, max(560, sw // 2)),
                         min(540, max(420, sh // 2)))
            root.resizable(True, True)
        except Exception:
            pass


    def _styles(self):
        s = ttk.Style(); s.theme_use("clam")
        s.configure("Run.TButton", background=RED, foreground=WHITE,
                    font=("Calibri", 11, "bold"), padding=(16, 9), borderwidth=0)
        s.map("Run.TButton", background=[("active", RED_DARK), ("disabled", "#aaa")])
        s.configure("Browse.TButton", background=NAVY, foreground=WHITE,
                    font=("Calibri", 10), padding=(10, 6), borderwidth=0)
        s.map("Browse.TButton", background=[("active", "#1a2550")])
        s.configure("Link.TButton", background=NAVY, foreground=WHITE,
                    font=("Calibri", 10), padding=(10, 6), borderwidth=0)
        s.map("Link.TButton", background=[("active", "#1a2550")])
        s.configure("Accent.Horizontal.TProgressbar",
                    troughcolor="#dde6f0", background=RED, borderwidth=0)


    def _extract_embedded(self, b64, filename):
        """Decode an embedded base64 asset into a temp file; return path or None."""
        try:
            if not b64:
                return None
            import base64 as _b64, tempfile, os
            target = os.path.join(tempfile.gettempdir(), filename)
            with open(target, "wb") as fh:
                fh.write(_b64.b64decode(b64))
            return target if os.path.isfile(target) else None
        except Exception:
            return None


    def _lock_header_colors(self, widget, navy):
        """Recursively bind <Enter>/<Leave> on all header widgets to force navy."""
        try:
            widget.bind("<Enter>", lambda e, w=widget, c=navy: w.configure(bg=c) if not isinstance(w, type(None)) else None)
            widget.bind("<Leave>", lambda e, w=widget, c=navy: w.configure(bg=c) if not isinstance(w, type(None)) else None)
        except Exception:
            pass
        try:
            for child in widget.winfo_children():
                self._lock_header_colors(child, navy)
        except Exception:
            pass
    def _header(self):
        self.header_mgr = FixedHeaderManager(self.root, title="GFH Inventory Aging Processor")
        hdr = tk.Frame(self.root, bg=NAVY, height=108)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        hdr._tag = "header"

        hdr.bind("<Enter>", lambda e, w=hdr: w.configure(bg=NAVY))
        hdr.bind("<Leave>", lambda e, w=hdr: w.configure(bg=NAVY))
        self._wordmark_img = None
        if _HAS_PIL:
            png_path = os.path.join(get_script_dir(), LOGO_PNG_NAME)
            try:
                if os.path.exists(png_path):
                    img = _PI.open(png_path).convert("RGBA")
                else:
                    data = base64.b64decode(HEADER_LOGO_B64.replace("\n", "").replace(" ", "").strip())
                    img = _PI.open(io.BytesIO(data)).convert("RGBA")
                bg2 = _PI.new("RGBA", img.size, (9, 13, 38, 255))
                bg2.paste(img, mask=img.split()[3])
                img = bg2.convert("RGB"); img.thumbnail((260, 82), _PI.Resampling.LANCZOS)
                self._wordmark_img = _PIT.PhotoImage(img)
            except Exception:
                pass
        lf = tk.Frame(hdr, bg=NAVY); lf.place(relx=0, rely=0.5, anchor="w", x=24)
        lf._tag = "header"
        if self._wordmark_img:
            tk.Label(lf, image=self._wordmark_img, bg=NAVY).pack()
        else:
            tk.Label(lf, text="GFH TELECOM", font=("Calibri", 16, "bold"), fg=RED, bg=NAVY).pack()
        tf = tk.Frame(hdr, bg=NAVY); tf.place(relx=0.5, rely=0.5, anchor="center")
        tf._tag = "header"
        tk.Label(tf, text="INVENTORY AGING PROCESSOR",
                 font=("Calibri", 18, "bold"), fg=WHITE, bg=NAVY).pack()
        tk.Label(tf, text="Process & Email  |  Upload to Google Sheets  |  Executive Dashboard",
                 font=("Calibri", 9), fg=WHITE, bg=NAVY).pack()

        theme_btn = create_theme_toggle_button(hdr, self.theme_manager, on_toggle=self._apply_theme)
        theme_btn.place(relx=0.98, rely=0.5, anchor="e")

        self._lock_header_colors(hdr, NAVY)

        self._lock_header_colors(hdr, NAVY)

    def _apply_theme(self, colors=None):
        apply_theme_to_window(self.root, self.theme_manager)

    def _body(self):
        body = tk.Frame(self.root, bg=LIGHT)
        body.pack(fill="both", expand=True, padx=24, pady=18)

        # Drop zone
        outer = tk.Frame(body, bg="#b8cce4", bd=0)
        outer.pack(fill="x", pady=(0, 14))
        self.drop_frame = tk.Frame(outer, bg=WHITE); self.drop_frame.pack(fill="x", padx=2, pady=2)
        dnd_note = "" if HAS_DND else "\n(install tkinterdnd2 for drag-and-drop)"
        self._dlbl = tk.Label(self.drop_frame,
            text=f"Drag & Drop  Inventory_Aging.xlsx  here\nor click Browse to select{dnd_note}",
            font=("Calibri", 11), fg="#4a6080", bg=WHITE, pady=28, cursor="hand2")
        self._dlbl.pack(fill="x")
        self._dlbl.bind("<Button-1>", lambda e: self._browse())
        if HAS_DND:
            for w in (self.drop_frame, self._dlbl):
                w.drop_target_register(DND_FILES)
                w.dnd_bind("<<Drop>>", self._on_drop)

        # File row
        row = tk.Frame(body, bg=LIGHT); row.pack(fill="x", pady=(0, 14))
        row.columnconfigure(0, weight=1)
        tk.Entry(row, textvariable=self.file_path, state="readonly",
                 font=("Calibri", 9), relief="flat", bg="#e8eff8", fg=NAVY,
                 readonlybackground="#e8eff8",
                 highlightbackground="#b0c4de", highlightthickness=1
                 ).grid(row=0, column=0, sticky="ew", ipady=5, padx=(0, 8))
        ttk.Button(row, text="Browse", style="Browse.TButton",
                   command=self._browse).grid(row=0, column=1)

        # Options
        opt = tk.Frame(body, bg=LIGHT); opt.pack(fill="x", pady=(0, 14))
        self.send_var = tk.BooleanVar(value=True)
        tk.Checkbutton(opt, text="Send emails via Outlook Desktop after processing",
                       variable=self.send_var, bg=LIGHT, fg=NAVY,
                       activebackground=LIGHT, selectcolor=WHITE,
                       font=("Calibri", 10)).pack(side="left")

        # Action buttons
        btn_row = tk.Frame(body, bg=LIGHT); btn_row.pack(fill="x", pady=(0, 12))
        self.run_btn = ttk.Button(btn_row, text="Process & Email Reports",
                                  style="Run.TButton", command=self._start_local)
        self.run_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))

        gs_state = "!disabled" if (_HAS_GSPREAD and _HAS_GSPREAD_FMT) else "disabled"
        self.upload_btn = ttk.Button(btn_row, text="Upload to Google Sheets",
                                     style="Run.TButton", command=self._start_google,
                                     state=gs_state)
        self.upload_btn.pack(side="left", fill="x", expand=True, padx=(4, 4))

        self.link_btn = ttk.Button(btn_row, text="Open Google Sheet",
                                   style="Link.TButton", command=self._open_sheet,
                                   state="disabled")
        self.link_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        self.settings_btn = ttk.Button(btn_row, text="Settings (Contacts)",
                                        style="Browse.TButton", command=self._open_settings)
        self.settings_btn.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # Progress
        self.progress = ttk.Progressbar(body, mode="indeterminate",
                                      style="Accent.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(0, 10))

        # Log
        tk.Label(body, text="Activity Log", font=("Calibri", 9, "bold"),
                 fg=NAVY, bg=LIGHT).pack(anchor="w")
        self.log_box = scrolledtext.ScrolledText(
            body, height=11, font=("Consolas", 8),
            bg="#10182e", fg="#a8d8ff", relief="flat", state="disabled", wrap="word")
        self.log_box.pack(fill="both", expand=True)

    def _open_settings(self):
        SettingsDialog(self.root)

    def _copyright_bar(self):
        bar = tk.Frame(self.root, bg=NAVY, height=26)
        bar.pack(fill="x", side="bottom"); bar.pack_propagate(False)
        tk.Label(bar, text=COPYRIGHT_TEXT, bg=NAVY, fg="#9d9db8",
                 font=("Calibri", 8)).pack(pady=4)

    def _on_drop(self, event):
        p = event.data.strip().strip("{}")
        if p.lower().endswith((".xlsx", ".xlsm", ".xls")):
            self.file_path.set(p); self._dlbl.config(fg="#1a6630")
            self._log(f"File dropped: {Path(p).name}")
        else:
            self._log("Please drop an Excel (.xlsx) file.")

    def _browse(self):
        p = filedialog.askopenfilename(title="Select Inventory Aging Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")])
        if p: self.file_path.set(p); self._log(f"Selected: {Path(p).name}")

    def _log(self, msg):
        self._q.put(msg)

    def _poll(self):
        try:
            while True:
                msg = self._q.get_nowait()
                self.log_box.config(state="normal")
                self.log_box.insert("end", msg + "\n"); self.log_box.see("end")
                self.log_box.config(state="disabled")
        except queue.Empty:
            pass
        self.root.after(80, self._poll)

    def _start_local(self):
        if self._running: return
        p = self.file_path.get().strip()
        if not p or not Path(p).exists():
            self._log("ERROR: Please select a valid Excel file first."); return
        self._running = True
        self.run_btn.config(state="disabled"); self.upload_btn.config(state="disabled")
        self.progress.start(12)
        self._log("-" * 55)
        _cfg = load_contacts_config()
        _sender_email, _cc_full, _ = contacts_config_to_legacy(_cfg)
        self._log(f"Input : {Path(p).name}")
        self._log(f"Sender: {_sender_email or '(not set — see Settings)'}  |  CC: {_cc_full or '(none)'}")
        self._log(f"Emails: {'YES' if self.send_var.get() else 'NO (disabled)'}")
        self._log("-" * 55)
        threading.Thread(target=self._worker_local,
                         args=(Path(p), self.send_var.get()), daemon=True).start()

    def _worker_local(self, input_path, send_email):
        try:
            run_local_pipeline(input_path, self._log, send_email=send_email)
            self.root.after(0, self._done)
        except Exception as exc:
            self._log(f"ERROR: {exc}"); self._log(traceback.format_exc())
            self.root.after(0, self._done)

    def _start_google(self):
        if self._running: return
        p = self.file_path.get().strip()
        if not p or not Path(p).exists():
            self._log("ERROR: Please select a valid Excel file first."); return
        self._running = True
        self.run_btn.config(state="disabled"); self.upload_btn.config(state="disabled")
        self.progress.start(12)
        self._log("-" * 55)
        self._log(f"Input: {Path(p).name}")
        self._log("Uploading to Google Sheets...")
        self._log("-" * 55)
        threading.Thread(target=self._worker_google, args=(p,), daemon=True).start()

    def _worker_google(self, excel_path):
        try:
            url = run_google_pipeline(excel_path, self._log)
            self.sheet_url_cache = url
            self.root.after(0, self._done_google, url)
        except PipelineError as e:
            self._log(f"ERROR: {e}")
            self.root.after(0, self._done)
        except Exception as e:
            self._log(f"ERROR: {e}"); self._log(traceback.format_exc())
            self.root.after(0, self._done)

    def _done(self):
        self.progress.stop()
        self.run_btn.config(state="normal")
        gs_state = "normal" if (_HAS_GSPREAD and _HAS_GSPREAD_FMT) else "disabled"
        self.upload_btn.config(state=gs_state)
        self._running = False

    def _done_google(self, url):
        self._done()
        self.link_btn.config(state="normal")
        messagebox.showinfo(APP_TITLE, f"Upload complete!\n\nGoogle Sheet:\n{url}")

    def _open_sheet(self):
        if self.sheet_url_cache:
            webbrowser.open(self.sheet_url_cache)


# ==========================================================
# ENTRY POINT
# ==========================================================
def _enable_dpi_awareness() -> None:
    """Make Windows report physical pixels so winfo_screen* is accurate on
    high-DPI displays (1080p, 1440p, 2K, 4K, DPI-scaled laptops)."""
    if sys.platform != "win32":
        return
    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)  # system DPI aware
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass


def main():
    _enable_dpi_awareness()
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=None)
    parser.add_argument("--no-email", action="store_true")
    parser.add_argument("--no-upload", action="store_true")
    args = parser.parse_args()

    if args.input:
        input_path = Path(args.input)

        saved = run_local_pipeline(input_path, print, send_email=not args.no_email)
        for dist, fp in saved:
            print(f"Saved: {fp.name}")

        if not args.no_upload and _HAS_GSPREAD and _HAS_GSPREAD_FMT:
            url = run_google_pipeline(str(input_path), print)
            print(f"Google Sheet: {url}")
        return

    root = (TkinterDnD.Tk() if HAS_DND else tk.Tk())
    App(root); root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        pass
    except Exception:
        traceback.print_exc()
        try:
            _r = tk.Tk(); _r.withdraw()
            messagebox.showerror("Fatal Error", traceback.format_exc())
        except:
            pass
